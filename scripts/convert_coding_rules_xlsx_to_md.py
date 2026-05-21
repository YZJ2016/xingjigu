from __future__ import annotations

from datetime import datetime
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


SOURCE = Path("/Users/yuzaijiang/Downloads/【TOPS-R】コーディング規約.xlsx")
OUTPUT = Path("docs/TOPS-R_coding_rules.md")


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    return text.strip()


def escape_markdown_table(text: str) -> str:
    return (
        text.replace("\\", "\\\\")
        .replace("|", "\\|")
        .replace("\n", "<br>")
    )


def merged_child_cells(sheet: Worksheet) -> set[tuple[int, int]]:
    children: set[tuple[int, int]] = set()
    for merged_range in sheet.merged_cells.ranges:
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if row != merged_range.min_row or col != merged_range.min_col:
                    children.add((row, col))
    return children


def sheet_matrix(value_sheet: Worksheet, formula_sheet: Worksheet) -> list[list[str]]:
    merged_children = merged_child_cells(value_sheet)
    rows: list[list[str]] = []
    max_row = max(value_sheet.max_row, formula_sheet.max_row)
    max_column = max(value_sheet.max_column, formula_sheet.max_column)
    for row in range(1, max_row + 1):
        values: list[str] = []
        for col in range(1, max_column + 1):
            if (row, col) in merged_children:
                values.append("")
                continue
            text = cell_text(value_sheet.cell(row, col).value)
            if not text:
                text = cell_text(formula_sheet.cell(row, col).value)
            values.append(text)
        rows.append(values)
    return rows


def non_empty_columns(rows: list[list[str]]) -> list[int]:
    cols: set[int] = set()
    for row in rows:
        for index, value in enumerate(row):
            if value:
                cols.add(index)
    return sorted(cols)


def is_rule_heading(row: list[str]) -> bool:
    candidates = row[:4]
    for index, value in enumerate(candidates):
        if not value:
            continue
        if index == 0 and (value.rstrip(".").isdigit() or value.endswith(".")):
            return True
        if re.match(r"^\d+[\.\uff0e]", value):
            return True
        if value.startswith("第") or value.startswith("Tips"):
            return True
    return False


def heading_text(row: list[str]) -> str:
    for value in row[:4]:
        if value and not value.rstrip(".").isdigit() and value != ".":
            return value
    return next((value for value in row if value), "その他")


def compact_table(rows: list[list[str]], columns: list[int]) -> list[list[str]]:
    result: list[list[str]] = []
    for row in rows:
        selected = [row[col] if col < len(row) else "" for col in columns]
        if not any(selected):
            continue
        result.append(selected)
    return result


def write_markdown_table(lines: list[str], headers: list[str], rows: list[list[str]]) -> None:
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for row in rows:
        lines.append("| " + " | ".join(escape_markdown_table(value) for value in row) + " |")
    lines.append("")


def render_sheet(value_sheet: Worksheet, formula_sheet: Worksheet) -> list[str]:
    matrix = sheet_matrix(value_sheet, formula_sheet)
    columns = non_empty_columns(matrix)
    title = value_sheet.title
    lines: list[str] = [f"# {title}", ""]

    if not columns:
        lines.append("_空シート_")
        lines.append("")
        return lines

    current_rule = "その他"
    buffer: list[list[str]] = []

    def flush() -> None:
        nonlocal buffer
        if buffer:
            block_columns = non_empty_columns(buffer)
            headers = ["行"] + [value_sheet.cell(1, col + 1).column_letter for col in block_columns[1:]]
            rows = compact_table(buffer, block_columns)
            if rows:
                write_markdown_table(lines, headers, rows)
        buffer = []

    if title in {"表紙", "改訂記録"}:
        lines.append(f"## {current_rule}")
        lines.append("")

    for row_index, row in enumerate(matrix, start=1):
        if title not in {"表紙", "改訂記録"} and is_rule_heading(row):
            flush()
            current_rule = heading_text(row)
            lines.append(f"## {current_rule}")
            lines.append("")
        if any(row[col] if col < len(row) else "" for col in columns):
            buffer.append([str(row_index)] + [row[col] if col < len(row) else "" for col in columns])

    if not any(line.startswith("## ") for line in lines):
        lines.append(f"## {current_rule}")
        lines.append("")
    flush()
    return lines


def main() -> None:
    values_workbook = load_workbook(SOURCE, data_only=True, read_only=False)
    formula_workbook = load_workbook(SOURCE, data_only=False, read_only=False)
    lines = [
        "# 【TOPS-R】コーディング規約",
        "",
        f"変換元: `{SOURCE.name}`",
        "",
    ]
    for value_sheet in values_workbook.worksheets:
        lines.extend(render_sheet(value_sheet, formula_workbook[value_sheet.title]))
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    print(OUTPUT.resolve())


if __name__ == "__main__":
    main()
